#!/usr/bin/env python
# coding: utf-8

# In[1]:


from openpyxl import Workbook
import glob


# In[2]:


#declare
class CMD:
    def __init__(self):
        self.content = []
        self.time = ''
        self.index = ''
        self.distance = ''
class remain_time:
    def __init__(self):
        self.remaining = ''
        self.time = ''
        self.target_phase = ''
        self.content = []
        self.distance = []
class service:
    def __init__(self):
        self.content = []
        self.start_time = ''
        self.end_time = ''
        self.index = ''
class STATUS:
    def __init__(self):
        self.time = ''
        self.content = ''


# In[3]:


dirpathPattern = '*.log'
date =''
result = glob.glob(dirpathPattern)


# In[4]:


tsp_list = []
evsp_list = []
tsp_start_time = []
tsp_end_time = []
evsp_start_time = []
evsp_end_time = []
tsp_remain_list = []
tsp_cmd_list = []
tsp_content = []
evsp_cmd_list = []
evsp_remain_list = []
evsp_status_list = []
evsp_content = []


# In[5]:


def read_logfile(r):
    global date
    date = r.split(' ')[0]
    log_list = []
    log_file = open(r)
    date = r.split(' ')[0]
    for log in log_file:
        log_list.append(str(log))
    return log_list


# In[6]:


def segmentation(log_list):
    start_tsp = 'TSP cloud packet rx: CMD(4)\n'
    end_tsp = 'TSP cloud packet rx: CMD(5)\n'
    start_evsp = 'EVSP OBU packet rx: ACTIVATE\n'
    end_evsp = 'EVSP OBU packet rx: TERMINATE\n'
    tsp_list = []
    evsp_list = []
    global tsp_start_time
    global tsp_end_time
    global evsp_start_time
    global evsp_end_time
    tsp_start_time = []
    tsp_end_time = []
    tsp = service()
    evsp = service()
    #TSP
    flag = 0
    for i, log in enumerate(log_list):
        if log == start_tsp and flag == 0:
            tsp.start_time = log_list[i-3].strip("\n")#.split(' ')[1]
            tsp_start_time.append(log_list[i-3].strip("\n"))
            #print(tsp.start_time)
            flag = 1
        if flag == 1 and log != end_tsp:
            tsp.content.append(log)
        elif log == end_tsp:
            tsp.end_time = log_list[i-3].strip('\n')
            tsp_end_time.append(log_list[i-3].strip("\n"))
            tsp_list.append(tsp)
            tsp = service()  
            flag = 0
    #EVSP
    flag = 0
    evsp_cmd = CMD()
    evsp_cmd_list = []
    for i, log in enumerate(log_list):
        if log == start_evsp:
            evsp.start_time = log_list[i-1].split(' ')[1].strip('\n')#.split(' ')[1]
            evsp_start_time.append(log_list[i-1].split(' ')[1].strip('\n'))
            evsp_cmd.time = log_list[i-1].split(' ')[1].strip('\n')
            evsp_cmd.index = int(log_list[i-6].split('(')[1].strip('\n').strip(')'))
            index = 0
            while(1):
                if log_list[i-5+index].split("[")[0] == 'cmd':
                    evsp_cmd.content.append(log_list[i-5+index])
                else:
                    break
                index += 1
            evsp_cmd_list.append(evsp_cmd)
            evsp_cmd = CMD()
            flag = 1
        if flag == 1 and log != end_evsp:
            evsp.content.append(log)
        elif log == end_evsp and flag ==1:
            evsp.end_time = log_list[i-1].split(' ')[1].strip('\n')
            evsp_end_time.append(log_list[i-1].split(' ')[1].strip('\n'))
            evsp_cmd.time = log_list[i-1].split(' ')[1].strip('\n')
            evsp_cmd.index = int(log_list[i-6].split('(')[1].strip('\n').strip(')'))
            index = 0
            while(1):
                if log_list[i-5+index].split("[")[0] == 'cmd':
                    evsp_cmd.content.append(log_list[i-5+index])
                else:
                    break
                index += 1
            evsp_cmd_list.append(evsp_cmd)
            evsp_cmd = CMD()
            evsp_list.append(evsp)
            evsp = service()  
            flag = 0 
    global date
    flag = 0
    evsp_remain = remain_time()
    evsp_remain_list = []
    for i, log in enumerate(log_list):
        if log == start_evsp and flag == 0:
            evsp_remain.time = log_list[i-1].split(' ')[1].strip('\n')
            flag = 1
        if flag == 1 and log.split(":")[0] == 'target phase':
            evsp_remain.target_phase = log.strip('\n')
        if flag == 1 and log.split(":")[0] == 'cycle':
            evsp_remain.content.append(log)
        if flag ==1 and log.split(' ')[0] == date:
            evsp_remain_list.append(evsp_remain)
            evsp_remain = remain_time()
            flag = 0
    flag = 0
    evsp_remain = remain_time()
    for i, log in enumerate(log_list):
        if log == end_evsp and flag == 0:
            evsp_remain.time = log_list[i-1].split(' ')[1].strip('\n')
            flag = 1
        if flag == 1 and log.split(":")[0] == 'cycle':
            evsp_remain.content.append(log)
        if flag == 1 and log.split(' ')[0] == date:
            evsp_remain_list.append(evsp_remain)
            evsp_remain = remain_time()
            flag = 0
    flag = 0
    status = STATUS()
    status_list = []
    for i, log in enumerate(log_list):
        if log.split(':')[0] == 'command_buf_polling' and flag == 0:
            status.time = log_list[i-1].strip('\n')
            flag = 1
        if flag == 1 and log.split(":")[0] == 'current signal status':
            status.content = log.split(":")[1].strip('\n')
        if flag == 1 and log.split(" ")[0] == date:
            status_list.append(status)
            status = STATUS()
            flag = 0
    return tsp_list, evsp_list, evsp_cmd_list, evsp_remain_list, status_list


# In[7]:


def tsp_remain_time(tsp_list):
    flag = 0
    tsp_remain = remain_time()
    tsp_remain_list = []
    for tsp in tsp_list:
        flag = 0
        for i, item in enumerate(tsp.content):
            if item.split(':')[0] == 'TSP supermatrix lookup' and flag == 0:
                tsp_remain.time = tsp.content[i-1].strip('\n')
                flag = 1
            if flag == 1 and item.split(":")[0] == 'distance':
                tsp_remain.distance = item.strip('\n')
            if flag == 1 and item.split(":")[0] == 'remaining':
                tsp_remain.remaining = item.strip('\n')
            if flag == 1 and item.split(":")[0] == 'target phase':
                tsp_remain.target_phase = item.strip('\n')
            if flag == 1 and item.split(":")[0] == 'cycle':
                for index in range(16):
                    tsp_remain.content.append(tsp.content[i+index].strip('\n'))
                tsp_remain_list.append(tsp_remain)
                tsp_remain = remain_time()
                flag = 0
    t_list = []
    for t in tsp_remain_list:
        t_list.append(t.time)
    flag = 0
    cmd = CMD()
    cmd_list = []
    for tsp in tsp_list:
        flag = 0
        for i, item in enumerate(tsp.content):
            if item.split(':')[0] == 'command buffer' and tsp.content[i+6].split(':')[0] == 'TSP supermatrix lookup':
                cmd.time = tsp.content[i-1].strip('\n')
                cmd.index = int(item.split('(')[1].strip('\n').strip(')'))
                flag = 1
            if flag == 1 and item.split("[")[0] == 'cmd':
                cmd.content.append(item)   
            if flag == 1 and item.split(" ")[0] == date:
                cmd_list.append(cmd)
                cmd = CMD()
                flag = 0
    return tsp_remain_list, cmd_list


# In[8]:


def make_content(remain_list, cmd_list, tsp_list):
    content = []
    t_list = []
    global tsp_start_time
    for c in cmd_list:
        t_list.append(c.time)
    for remain in remain_list:
        right_cmd = ''
        cycle = ''
        phase = ''
        if remain.time in t_list:
            for c in remain.content:
                if c.split(' ')[6] == '(0)':
                    cycle = c.strip("\n")
                    phase = c.split(' ')[3].strip(',').strip('\n')
                    break
            for j in cmd_list:
                if j.time == remain.time:
                    x = str((j.index + int(cycle.split(' ')[1].strip(',')))%2)
                    y = phase
                    for c in j.content:
                        if c.split('[')[1].strip(']') == x and c.split(']')[1].strip('[') == y:
                            right_cmd = c.strip('\n')
                            content.append(['', '', remain.time.split(' ')[1], remain.distance, remain.remaining, remain.target_phase, cycle, right_cmd])
                            break    
        else:
            content.append(['', '', remain.time.split(' ')[1], remain.distance, remain.remaining, remain.target_phase, '', ''])
    for start in tsp_start_time:
        for i, c in enumerate(content):
            if c[2] == start.split(' ')[1]:
                c[0] = start.split(' ')[1]
                c[1] = 'CMD(4)'
                break
            if i == len(content)-1:
                content.append([start.split(' ')[1].strip('\n'), 'CMD(4)', start.split(' ')[1].strip('\n'), '', '','','',''])
    for tsp in tsp_list:
        for i, c in enumerate(content):
            if c[2] == tsp.end_time.split(' ')[1]:
                c[0] = tsp.end_time.split(' ')[1]
                c[1] = 'CMD(5)'
                break
            if i == len(content)-1:
                content.append([tsp.end_time.split(' ')[1].strip('\n'), 'CMD(5)', tsp.end_time.split(' ')[1].strip('\n'), '', '','','',''])
    return content
            


# In[9]:


def make_evsp_content(remain_list, status_list, cmd_list, evsp_list):
    content = []
    line = ['','','','','','']
    global evsp_start_time
    global evsp_end_time
    check = evsp_start_time
    for i, start in enumerate(evsp_start_time):
        for status in status_list:
            if start == status.time.split(' ')[1] and (start in check):
                line = [start, 'ACTIVATE', status.content,'','','']
                content.append(line)
                line = ['','','','','','']
                check.remove(start)
            else:
                line = [status.time.split(' ')[1], '', status.content,'','','']
                content.append(line)
                line = ['','','','','',''] 
    for end in evsp_end_time:
        for c in content:
            if c[0] == end:
                c[1] = 'TERMINATE'
    cycle = []
    phase = []
    x = []
    y = []
    for c in content:
        for remain in remain_list:
            cycle = []
            phase = []
            x = []
            y = []
            if c[0] == remain.time and (c[1] == 'ACTIVATE' or c[1] == 'TERMINATE'):
                c[3] = remain.target_phase
                if len(remain.content) > 1:
                    c[4] = remain.content[0] + remain.content[1]
                else:
                    c[4] = remain.content[0]
                for cmd in cmd_list:
                    if cmd.time == remain.time:
                        for rc in remain.content:
                            cycle.append(rc.split(' ')[1].strip(','))
                            phase.append(rc.split(' ')[3].strip(',').strip('\n'))
                        #cycle1 = remain.content[0].split(' ')[1].strip(',')
                        #phase1 = remain.content[0].split(' ')[3].strip(',').strip('\n')
                        #if len(remain.content) > 1:
                        #    cycle2 = remain.content[1].split(' ')[1].strip(',')
                        #    phase2 = remain.content[1].split(' ')[3].strip(',').strip('\n')
                        #    even = 1
                        for i in range(len(cycle)):
                            x.append(str((cmd.index + int(cycle[i]))%2))
                            y.append(phase[i])
                        for con in cmd.content:
                            for i in range(len(x)):
                                if con.split('[')[1].strip(']') == x[i] and con.split(']')[1].strip('[') == y[i]:
                                    c[5] += con
    for i, c in enumerate(content):
        if content[i][1] == 'ACTIVATE'and content[i-1][1] == 'ACTIVATE':
            content[i-1] = [content[i-1][0], '', content[i-1][2],'','','']  
    for i, c in enumerate(content):
        if len(c[4].split('\n'))>2 and len(c[5].split('\n'))>2:# and (c[1] == 'ACTIVATE'or c[1] == 'TERMINATE'):
            tmp = c[4].split('\n')
            tmp2 = c[5].split('\n')
            c[4] = tmp[0].strip('\n')
            c[5] = tmp2[0].strip('\n')
            time = c[0]
            for j in range(1, len(tmp)-1):
                content.insert(i+j, [time,'','','',tmp[j].strip('\n'),tmp2[j].strip('\n')])
    return content
            


# In[10]:


def write_file(content, evsp_content, r):
    content = sorted(content, key = lambda x : x[2])
    #evsp_content = sorted(evsp_content, key = lambda x : x[0])     
    index = 0
    #TSP 
    wb = Workbook()
    sheet = wb.worksheets[0]
    sheet.title = 'tsp_log'
    for i, c in enumerate(content):
        if c[1] != 'CMD(5)':
            sheet.append(c)
        else:
            sheet.append(c)
            sheet = wb.create_sheet('tsp_log', index=i)
    #EVSP
    index = len(wb.worksheets)
    sheet = wb.worksheets[index-1]
    sheet.title = 'evsp_log'
    flag = 0
    for i, c in enumerate(evsp_content):
        if c[1] == 'ACTIVATE' and flag == 0:
            flag = 1
            sheet.append(c)    
        elif c[1] == 'TERMINATE'and flag == 1:
            sheet.append(c)
            sheet = wb.create_sheet('evsp_log', index=i+index)
            flag = 0
        elif flag == 1 and c[1] != 'ACTIVATE':
            sheet.append(c)       
    for worksheet in wb.worksheets:
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            # Since Openpyxl 2.6, the column name is  ".column_letter" as .column became the column number (1-based) 
            for cell in col:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[str(column)].width = adjusted_width
    wb.save(r.split('.')[0]+'.xlsx')


# In[11]:


#tsps' content storted in tsp_list
#evsps' content storted in evsp_list  
for r in result:
    tsp_list = []
    evsp_list = []
    tsp_start_time = []
    tsp_end_time = []
    evsp_start_time = []
    evsp_end_time = []
    tsp_remain_list = []
    tsp_cmd_list = []
    tsp_content = []
    evsp_cmd_list = []
    evsp_remain_list = []
    evsp_status_list = []
    evsp_content = []
    tsp_list, evsp_list, evsp_cmd_list, evsp_remain_list, evsp_status_list = segmentation(read_logfile(r))
    tsp_remain_list, tsp_cmd_list = tsp_remain_time(tsp_list)
    tsp_content = make_content(tsp_remain_list, tsp_cmd_list, tsp_list)
    evsp_content = make_evsp_content(evsp_remain_list, evsp_status_list, evsp_cmd_list, evsp_list)
    write_file(tsp_content, evsp_content, r)


# In[ ]:




